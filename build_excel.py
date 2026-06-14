import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.series import DataPoint
import os

BLUE_D  = "1F4E79"
BLUE_M  = "2E75B6"
BLUE_L  = "BDD7EE"
ORANGE  = "C55A11"
ORANGE_L= "FCE4D6"
GREEN_D = "375623"
GREEN_L = "E2EFDA"
RED     = "C00000"
GRAY_H  = "F2F2F2"
WHITE   = "FFFFFF"
GOLD    = "FFD700"

def hdr(ws, row, col, text, bg=BLUE_D, fg=WHITE, bold=True, size=11, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=bold, color=fg, size=size, name="Arial")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c

def val(ws, row, col, v, bold=False, color="000000", align="center", fmt=None, italic=False):
    c = ws.cell(row=row, column=col, value=v)
    c.font      = Font(bold=bold, color=color, size=10, name="Arial", italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt: c.number_format = fmt
    return c

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def col_border(ws, r1, r2, cols):
    for r in range(r1, r2+1):
        for c in cols:
            ws.cell(r, c).border = thin_border()

def kpi_box(ws, row, col, label, value, bg=BLUE_M):
    ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    c1 = ws.cell(row=row,   column=col, value=label)
    c2 = ws.cell(row=row+1, column=col, value=value)
    c1.font = Font(bold=True, color=WHITE, size=9, name="Arial")
    c2.font = Font(bold=True, color=WHITE, size=13, name="Arial")
    for c in [c1, c2]:
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

wb = Workbook()

# ══════════════════════════════════════════════════════
# TAB 1 — INVENTORY FORECAST
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Inventory Forecast"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 14
for c in ["B","C","D","E","F","G"]: ws1.column_dimensions[c].width = 18
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 20

ws1.merge_cells("A1:G1")
t = ws1["A1"]
t.value     = "PulseIndia Digital — Pre-Sales Inventory Forecast  |  Campaign Window: Oct 14 – Nov 25, 2024"
t.font      = Font(bold=True, color=WHITE, size=14, name="Arial")
t.fill      = PatternFill("solid", fgColor=BLUE_D)
t.alignment = Alignment(horizontal="center", vertical="center")

ws1.merge_cells("A2:G2")
s = ws1["A2"]
s.value     = "Holt-Winters Exponential Smoothing  |  Diwali Multiplier Applied (+22%)  |  15% Conservative Buffer"
s.font      = Font(italic=True, color="595959", size=9, name="Arial")
s.fill      = PatternFill("solid", fgColor=BLUE_L)
s.alignment = Alignment(horizontal="center", vertical="center")

# KPI Row
kpi_box(ws1, 4, 1, "Total Committable",   "15,560,975 impr",  BLUE_D)
kpi_box(ws1, 4, 3, "Total RFP Committed", "9,333,333 impr",   ORANGE)
kpi_box(ws1, 4, 5, "Net Buffer",          "6,227,642 impr",   GREEN_D)
kpi_box(ws1, 4, 7, "Overbooking Weeks",   "0  ✓ SAFE",        "375623")

headers = ["Week Start","Raw Forecast","Diwali-Adjusted","Committable\n(−15% Buffer)",
           "RFP-001\n(Weekly)","RFP-002\n(Weekly)","Total\nCommitted",
           "Net Available","Status"]
for i, h in enumerate(headers, 1):
    hdr(ws1, 7, i, h, wrap=True)
ws1.row_dimensions[7].height = 32

forecast = pd.read_csv("/home/claude/pulseIndia/outputs/forecast_results.csv")
for ri, row in forecast.iterrows():
    r = ri + 8
    status_ok = str(row["status"]).startswith("SAFE")
    row_bg    = GREEN_L if status_ok else "FFCCCC"
    vals = [
        row["week_start"],
        int(row["raw_forecast"]),
        int(row["diwali_adjusted"]),
        int(row["committable_after_buffer"]),
        int(row["rfp1_committed_weekly"]),
        int(row["rfp2_committed_weekly"]),
        int(row["total_committed"]),
        int(row["net_available_impressions"]),
        str(row["status"]),
    ]
    for ci, v in enumerate(vals, 1):
        c = val(ws1, r, ci, v)
        c.fill   = PatternFill("solid", fgColor=row_bg)
        c.border = thin_border()
        if ci in [2,3,4,5,6,7,8]:
            c.number_format = "#,##0"

ws1.row_dimensions[8].height  = 18
ws1.freeze_panes = "A8"

# Insert chart image
img1 = XLImage("/home/claude/pulseIndia/charts/chart3_forecast_vs_commitments.png")
img1.width, img1.height = 700, 320
ws1.add_image(img1, "A16")

# ══════════════════════════════════════════════════════
# TAB 2 — SELL-THROUGH ANALYSIS
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Sell-Through Analysis")
ws2.sheet_view.showGridLines = False
for c in ["A","B","C","D","E","F"]: ws2.column_dimensions[c].width = 18
ws2.row_dimensions[1].height = 40

ws2.merge_cells("A1:F1")
t2 = ws2["A1"]
t2.value     = "PulseIndia Digital — 16-Week Sell-Through Rate & CPM Analysis"
t2.font      = Font(bold=True, color=WHITE, size=14, name="Arial")
t2.fill      = PatternFill("solid", fgColor=BLUE_D)
t2.alignment = Alignment(horizontal="center", vertical="center")

history = pd.read_csv("/home/claude/pulseIndia/outputs/weekly_history.csv")
history["str_pct"] = history["sell_through"] * 100

kpi_box(ws2, 3, 1, "Avg Sell-Through", f"{history['str_pct'].mean():.1f}%", BLUE_D)
kpi_box(ws2, 3, 3, "Peak Week",        f"{history['str_pct'].max():.1f}%",  GREEN_D)
kpi_box(ws2, 3, 5, "Lowest Week",      f"{history['str_pct'].min():.1f}%",  RED)

hdrs2 = ["Week Start","Total Available","Total Served","Sell-Through %","Avg CPM (₹)","Revenue (₹)"]
for i, h in enumerate(hdrs2, 1):
    hdr(ws2, 6, i, h)

for ri, row in history.iterrows():
    r     = ri + 7
    str_v = row["str_pct"]
    bg    = GREEN_L if str_v >= 78 else (ORANGE_L if str_v >= 65 else "FFCCCC")
    data  = [row["week_start"], int(row["avail"]), int(row["served"]),
             str_v/100, row["avg_cpm"], row["revenue"]]
    fmts  = [None, "#,##0", "#,##0", "0.0%", "₹#,##0.00", "₹#,##0"]
    for ci, (v, f) in enumerate(zip(data, fmts), 1):
        c = val(ws2, r, ci, v)
        c.fill   = PatternFill("solid", fgColor=bg)
        c.border = thin_border()
        if f: c.number_format = f

img2 = XLImage("/home/claude/pulseIndia/charts/chart1_sell_through_trend.png")
img2.width, img2.height = 720, 310
ws2.add_image(img2, "A25")

img2b = XLImage("/home/claude/pulseIndia/charts/chart2_cpm_vs_str_scatter.png")
img2b.width, img2b.height = 500, 330
ws2.add_image(img2b, "A42")

# ══════════════════════════════════════════════════════
# TAB 3 — AUDIENCE SEGMENT HEATMAP
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Audience Segments")
ws3.sheet_view.showGridLines = False
for c in ["A","B","C","D","E","F","G"]: ws3.column_dimensions[c].width = 20

ws3.merge_cells("A1:G1")
t3 = ws3["A1"]
t3.value     = "PulseIndia Digital — Audience Segment Performance & RFP Targeting Validation"
t3.font      = Font(bold=True, color=WHITE, size=14, name="Arial")
t3.fill      = PatternFill("solid", fgColor=BLUE_D)
t3.alignment = Alignment(horizontal="center", vertical="center")

ws3.merge_cells("A2:G2")
n = ws3["A2"]
n.value     = "RFP-001 targets: Urban Male 25–44 | Mobile  •  RFP-002 targets: News Enthusiast | Desktop"
n.font      = Font(italic=True, color=WHITE, size=10, name="Arial")
n.fill      = PatternFill("solid", fgColor=BLUE_M)
n.alignment = Alignment(horizontal="center", vertical="center")

hdrs3 = ["User Segment","Device Type","Total Available","Avg Sell-Through %","Avg CPM (₹)","Total Revenue (₹)","RFP Match"]
for i, h in enumerate(hdrs3, 1):
    hdr(ws3, 4, i, h)

seg = pd.read_csv("/home/claude/pulseIndia/outputs/segment_analysis.csv")
rfp_matches = {
    ("urban_male_25_44","mobile"):   "✓ RFP-001",
    ("news_enthusiast","desktop"):   "✓ RFP-002",
}

for ri, row in seg.iterrows():
    r      = ri + 5
    match  = rfp_matches.get((row["user_segment"], row["device_type"]), "—")
    is_rfp = match != "—"
    bg     = GOLD if is_rfp else (GREEN_L if row["avg_str_pct"] >= 78 else GRAY_H)
    data   = [row["user_segment"].replace("_"," ").title(), row["device_type"].capitalize(),
              int(row["total_avail"]), row["avg_str_pct"]/100, row["avg_cpm"],
              int(row["total_avail"] * row["avg_str_pct"]/100 * row["avg_cpm"]/1000), match]
    fmts   = [None, None, "#,##0", "0.0%", "₹#,##0.00", "₹#,##0", None]
    for ci, (v, f) in enumerate(zip(data, fmts), 1):
        c = val(ws3, r, ci, v)
        c.fill   = PatternFill("solid", fgColor=bg)
        c.border = thin_border()
        if f: c.number_format = f
        if is_rfp: c.font = Font(bold=True, color="1F4E79", size=10, name="Arial")

img3 = XLImage("/home/claude/pulseIndia/charts/chart5_segment_heatmap.png")
img3.width, img3.height = 560, 340
ws3.add_image(img3, "A22")

# ══════════════════════════════════════════════════════
# TAB 4 — PMP vs OPEN AUCTION MODEL
# ══════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. PMP Revenue Model")
ws4.sheet_view.showGridLines = False
for c in ["A","B","C","D","E","F","G"]: ws4.column_dimensions[c].width = 22

ws4.merge_cells("A1:G1")
t4 = ws4["A1"]
t4.value     = "PulseIndia Digital — PMP vs Open Auction Revenue Scenario Model"
t4.font      = Font(bold=True, color=WHITE, size=14, name="Arial")
t4.fill      = PatternFill("solid", fgColor=BLUE_D)
t4.alignment = Alignment(horizontal="center", vertical="center")

# Assumptions box
ws4.merge_cells("A3:C3")
hdr(ws4, 3, 1, "MODEL ASSUMPTIONS", bg=BLUE_M, size=10)
assumptions = [
    ("Historical Avg CPM (INR)",      "₹88.85"),
    ("Historical Avg Sell-Through",   "76.0%"),
    ("P75 CPM — PMP Floor (INR)",     "₹89.37"),
    ("Total Committable Impressions",  "15,560,975"),
    ("Pricing Basis",                  "P75 CPM — advertisers pay premium for guarantee"),
    ("Source",                         "16-week historical ad server log (IVT-excluded)"),
]
for i, (k, v) in enumerate(assumptions, 4):
    ws4.cell(i, 1, k).font  = Font(bold=True, size=9, color="1F4E79", name="Arial")
    ws4.cell(i, 2, v).font  = Font(size=9, color="000000", name="Arial")
    ws4.cell(i, 1).fill = ws4.cell(i, 2).fill = PatternFill("solid", fgColor=BLUE_L)

hdrs4 = ["Scenario","Fill Rate","PMP CPM Floor (₹)","Open Auction Rev (₹)","PMP Revenue (₹)","Uplift (₹)","Uplift (%)"]
for i, h in enumerate(hdrs4, 1):
    hdr(ws4, 12, i, h)

pmp = pd.read_csv("/home/claude/pulseIndia/outputs/pmp_model.csv")
for ri, row in pmp.iterrows():
    r       = ri + 13
    uplift  = row["Revenue Uplift (%)"]
    bg      = GREEN_L if uplift > 0 else "FFCCCC"
    open_r  = row["Open Auction Revenue (INR)"]
    pmp_r   = row["PMP Revenue (INR)"]
    data    = [row["Scenario"], row["Fill Rate"], row["PMP CPM Floor (INR)"],
               open_r, pmp_r, pmp_r - open_r, uplift/100]
    fmts    = [None, None, "₹#,##0.00", "₹#,##0", "₹#,##0", "₹#,##0", "+0.0%;-0.0%"]
    for ci, (v, f) in enumerate(zip(data, fmts), 1):
        c = val(ws4, r, ci, v)
        c.fill   = PatternFill("solid", fgColor=bg)
        c.border = thin_border()
        if f: c.number_format = f

ws4.merge_cells("A17:G17")
rec = ws4["A17"]
rec.value = "★  RECOMMENDATION: PMP deal at ₹89.37 CPM floor (Base scenario, 85% fill) projects ₹1,18,205 revenue vs ₹1,05,107 open auction — a 12.5% uplift. Recommend PMP for both RFPs."
rec.font  = Font(bold=True, color=WHITE, size=10, name="Arial")
rec.fill  = PatternFill("solid", fgColor=GREEN_D)
rec.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws4.row_dimensions[17].height = 36

img4 = XLImage("/home/claude/pulseIndia/charts/chart4_pmp_vs_open_auction.png")
img4.width, img4.height = 560, 340
ws4.add_image(img4, "A19")

img4b = XLImage("/home/claude/pulseIndia/charts/chart6_ivt_analysis.png")
img4b.width, img4b.height = 640, 300
ws4.add_image(img4b, "A37")

out = "/home/claude/pulseIndia/outputs/PulseIndia_PreSales_InventoryReport.xlsx"
wb.save(out)
print("✓ Excel workbook saved")
